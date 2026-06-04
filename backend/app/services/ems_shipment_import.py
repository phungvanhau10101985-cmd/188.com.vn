from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime, timezone
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import Session

from app import crud
from app.models.order import Order, OrderStatus
from app.models.order_shipment import (
    EmsShippingImportBatch,
    EmsShippingImportBatchRow,
    EmsShippingRecord,
    OrderShipmentEvent,
)
from app.services import ems_tracking as ems_tracking_svc
from app.services import order_shipment_timeline as shipment_svc

_ORDER_CODE_RE = re.compile(r"\b(DH\d+)\b", re.IGNORECASE)
_DC_CODE_RE = re.compile(r"\b(DC\d+)\b", re.IGNORECASE)
_MADONHANG_RE = re.compile(r"madonhang\s*(DH\d+|DC\d+)", re.IGNORECASE)
_INVALID_ORDER_PLACEHOLDERS = frozenset({
    "ZALO", "XALO", "FACEBOOK", "FB", "SHOPEE", "TIKTOK", "MDH", "NONE", "N/A", "NA",
})

# file gui ems.xlsx (shop export)
_SHOP_EXPORT_HEADERS: dict[str, tuple[str, ...]] = {
    "reference_code": ("MA_VAN_DON",),
    "product_name": ("TEN_SP",),
    "weight": ("TRONG_LUONG",),
    "customer_name": ("TEN_KH",),
    "address": ("DIA_CHI_KH",),
    "phone": ("SDT_KH", "SDT"),
    "cod_amount": ("COD",),
    "product_code": ("MA_SP", "MA_SAN_PHAM", "MA_SP_KHO", "SKU", "MA_HANG"),
    "order_code": ("DON_HANG", "MA_DON_HANG", "MADONHANG", "ORDER_CODE"),
}
# Cột H (0-based index 7) trên file gửi EMS chuẩn
_COL_H_INDEX = 7
_SHOP_EXPORT_DEFAULTS: dict[str, int] = {
    "reference_code": 0,
    "product_name": 1,
    "weight": 2,
    "customer_name": 3,
    "address": 4,
    "phone": 5,
    "cod_amount": 6,
    "product_code": 7,
    "order_code": 8,
}

# EMS export cũ (MA_DON_HANG / TEN_NGUOI_NHAN / TONG_TIEN_THU_HO)
_COL_REF_NAMES = ("MA_DON_HANG", "MA THAM CHIEU", "MA_THAM_CHIEU")
_COL_RECIPIENT_NAMES = ("TEN_NGUOI_NHAN", "TEN NGUOI NHAN")
_COL_COD_NAMES = (
    "TONG_TIEN_THU_HO",
    "TONG TIEN THU HO",
    "TIEN_THU_HO",
    "TIEN THU HO",
    "THU_HO",
)
_EMS_EXPORT_DEFAULTS: dict[str, int] = {
    "reference_code": 3,
    "product_code": 7,
    "recipient_label": 9,
    "cod_amount": 15,
}


def _norm_header(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip())
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d").replace("Đ", "D")
    text = text.upper()
    return re.sub(r"\s+", "_", text)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in ("nan", "none"):
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _is_placeholder_order_code(code: str) -> bool:
    text = (code or "").strip().upper()
    if not text:
        return True
    if text in _INVALID_ORDER_PLACEHOLDERS:
        return True
    if text.startswith("FB+") or text.endswith("MDH"):
        return True
    return False


def _normalize_external_order_code(code: str) -> Optional[str]:
    text = (code or "").strip().upper()
    if not text or _is_placeholder_order_code(text):
        return None
    if _ORDER_CODE_RE.fullmatch(text) or _DC_CODE_RE.fullmatch(text):
        return text
    dh = _ORDER_CODE_RE.search(text)
    if dh:
        return dh.group(1).upper()
    dc = _DC_CODE_RE.search(text)
    if dc:
        return dc.group(1).upper()
    return None


def looks_like_recipient_not_sku(text: Optional[str]) -> bool:
    """Phát hiện chuỗi tên/SĐT/địa chỉ khách — không phải MA_SP cột H."""
    raw = _cell_str(text)
    if not raw:
        return False
    if "—" in raw or " · " in raw:
        return True
    lower = raw.lower()
    if any(
        m in lower
        for m in (
            "đường",
            "phường",
            "quận",
            "huyện",
            "thành phố",
            "tỉnh",
            "tp ",
            "ngõ",
            "thôn",
            "xã ",
        )
    ):
        return True
    if len(raw) > 48 and "/" not in raw:
        return True
    return False


def extract_warehouse_sku_from_ems_label(text: Optional[str]) -> Optional[str]:
    """
    Mã kho từ cột H (MA_SP) / nhãn EMS: lấy chuỗi trước dấu «-» đầu tiên.
    Vd. B7796/41/2-XANH LAM-*https://... → B7796/41/2
    """
    raw = _cell_str(text)
    if not raw or looks_like_recipient_not_sku(raw):
        return None
    head = raw
    for sep in ("-", "–", "—", "−"):
        if sep in head:
            head = head.split(sep, 1)[0]
            break
    head = head.strip()
    if not head or len(head) < 2:
        return None
    if _is_placeholder_order_code(head):
        return None
    if looks_like_recipient_not_sku(head):
        return None
    return head


def warehouse_sku_from_col_h_cell(text: Optional[str]) -> Optional[str]:
    """Chuẩn hóa ô cột H (MA_SP) → mã kho trước dấu «-»."""
    return extract_warehouse_sku_from_ems_label(text)


_SKU_FRAGMENT_RE = re.compile(
    r"([A-Za-z]\d{2,}(?:/[A-Za-z0-9]{1,14})+)\s*[-–—]",
    re.IGNORECASE,
)


def extract_ma_sp_from_text_blob(text: str) -> tuple[str, str]:
    """Trích MA_SP từ một ô (cột H, TEN_SP, hoặc chuỗi lẫn tên SP)."""
    raw = _cell_str(text)
    if not raw or looks_like_recipient_not_sku(raw):
        return "", ""
    match = _SKU_FRAGMENT_RE.search(raw)
    if match:
        frag = match.group(1).strip()
        sku = warehouse_sku_from_col_h_cell(frag + "-x") or frag
        if sku and not looks_like_recipient_not_sku(sku):
            return match.group(0).strip(), sku
    # Cả ô là MA_SP (vd. B7796/41/2-XANH LAM-*url) — không cắt ở dấu - đầu trong tên dài
    if len(raw) <= 96 and "/" in raw:
        sku = warehouse_sku_from_col_h_cell(raw) or ""
        if sku and not looks_like_recipient_not_sku(sku) and len(sku) <= 48:
            return raw, sku
    return "", ""


def _append_cell_candidates(
    raws: list[str],
    *,
    row: tuple[Any, ...],
    row_alt: tuple[Any, ...] | None,
    indices: list[int],
) -> None:
    seen: set[str] = set()
    for idx in indices:
        if idx < 0:
            continue
        for source in (row, row_alt or ()):
            if idx >= len(source):
                continue
            val = source[idx]
            if val is None:
                continue
            text = _cell_str(val)
            if not text:
                s = str(val).strip()
                if s.startswith("="):
                    continue
                text = _cell_str(s)
            if text and text not in seen:
                seen.add(text)
                raws.append(text)


def read_ma_sp_from_excel_row(
    row: tuple[Any, ...],
    col_map: dict[str, int],
    *,
    row_alt: tuple[Any, ...] | None = None,
) -> tuple[str, str]:
    """
    Đọc MA_SP từ dòng Excel — cột header MA_SP, cột H (7), TEN_SP (1), quét cả dòng.
    row_alt: cùng dòng đọc data_only=False (ô công thức Excel chưa cache).
    """
    indices: list[int] = []
    for key in ("product_code", "product_name"):
        mapped = col_map.get(key)
        if mapped is not None:
            indices.append(int(mapped))
    for idx in (_COL_H_INDEX, 8, 1, 6, 2):
        if idx not in indices:
            indices.append(idx)

    raws: list[str] = []
    _append_cell_candidates(raws, row=row, row_alt=row_alt, indices=indices)

    for raw in raws:
        found_raw, sku = extract_ma_sp_from_text_blob(raw)
        if sku:
            return found_raw or raw, sku

    for cell in row:
        text = _cell_str(cell)
        if not text or len(text) > 220:
            continue
        found_raw, sku = extract_ma_sp_from_text_blob(text)
        if sku:
            return found_raw or text, sku

    return "", ""


def extract_order_code_from_recipient(text: Optional[str]) -> Optional[str]:
    raw = (text or "").strip()
    if not raw:
        return None
    madonhang = _MADONHANG_RE.search(raw)
    if madonhang:
        return madonhang.group(1).upper()
    match = _ORDER_CODE_RE.search(raw)
    if match:
        return match.group(1).upper()
    dc = _DC_CODE_RE.search(raw)
    return dc.group(1).upper() if dc else None


def _resolve_order_code(*, direct: Any = None, product_code: Any = None, recipient_label: Any = None) -> Optional[str]:
    direct_code = _normalize_external_order_code(_cell_str(direct))
    if direct_code:
        return direct_code
    for source in (product_code, recipient_label):
        parsed = extract_order_code_from_recipient(_cell_str(source))
        if parsed:
            return parsed
    return None


def _is_shop_order_code(order_code: Optional[str]) -> bool:
    code = (order_code or "").strip().upper()
    return bool(code and _ORDER_CODE_RE.fullmatch(code))


def _build_recipient_label(*, customer_name: str = "", phone: str = "", address: str = "", legacy: str = "") -> str:
    if legacy.strip():
        return legacy.strip()
    parts = [customer_name.strip(), phone.strip()]
    label = " · ".join(p for p in parts if p)
    if address.strip():
        return f"{label} — {address.strip()}" if label else address.strip()
    return label


def _isoformat_value(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    iso = getattr(value, "isoformat", None)
    if callable(iso):
        return iso()
    return str(value)


def _parse_cod_amount(value: Any) -> Optional[int]:
    """Parse cột tiền thu hộ (VND). 0 ₫ là hợp lệ (đơn không thu hộ); ô trống → None."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        return int(round(value)) if value >= 0 else None
    text = _cell_str(value)
    if not text:
        return None
    cleaned = re.sub(r"[^\d,.-]", "", text)
    if not cleaned:
        return None
    if cleaned.count(".") > 1 and "," not in cleaned:
        cleaned = cleaned.replace(".", "")
    elif cleaned.count(",") > 1 and "." not in cleaned:
        cleaned = cleaned.replace(",", "")
    else:
        cleaned = cleaned.replace(",", "")
    try:
        amt = float(cleaned)
    except ValueError:
        return None
    if amt < 0:
        return None
    return int(round(amt))


def _match_header_field(key: str, aliases: tuple[str, ...]) -> bool:
    return key in aliases


def _find_header_map(rows: list[tuple[Any, ...]]) -> tuple[int, str, dict[str, int]]:
    for idx, row in enumerate(rows[:10]):
        shop_map: dict[str, int] = {}
        ems_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            key = _norm_header(cell)
            for field, aliases in _SHOP_EXPORT_HEADERS.items():
                if _match_header_field(key, aliases):
                    shop_map[field] = col_idx
            if _match_header_field(key, _COL_REF_NAMES):
                ems_map["reference_code"] = col_idx
            if _match_header_field(key, _SHOP_EXPORT_HEADERS["product_code"]):
                ems_map["product_code"] = col_idx
            if _match_header_field(key, _COL_RECIPIENT_NAMES):
                ems_map["recipient_label"] = col_idx
            if _match_header_field(key, _COL_COD_NAMES):
                ems_map["cod_amount"] = col_idx

        if shop_map.get("reference_code") is not None and (
            shop_map.get("order_code") is not None or shop_map.get("product_code") is not None
        ):
            merged = {**_SHOP_EXPORT_DEFAULTS, **shop_map}
            return idx, "shop_export", merged
        if ems_map.get("reference_code") is not None and ems_map.get("recipient_label") is not None:
            merged = {**_EMS_EXPORT_DEFAULTS, **ems_map}
            return idx, "ems_export", merged

    return 0, "shop_export", dict(_SHOP_EXPORT_DEFAULTS)


def parse_ems_export_rows(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    bio = io.BytesIO(file_bytes)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        raw_rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    raw_rows_alt: list[tuple[Any, ...]] = []
    try:
        bio.seek(0)
        wb_alt = load_workbook(bio, read_only=True, data_only=False)
        try:
            raw_rows_alt = [tuple(row) for row in wb_alt.active.iter_rows(values_only=True)]
        finally:
            wb_alt.close()
    except Exception:
        raw_rows_alt = []

    if not raw_rows:
        return [], ["File Excel trống."]

    header_idx, file_format, col_map = _find_header_map(raw_rows)
    if header_idx > 0:
        warnings.append(f"Phát hiện dòng tiêu đề ở hàng {header_idx + 1}.")
    if file_format == "shop_export":
        warnings.append(
            "Định dạng file gửi EMS: cột A mã vận đơn, I đơn hàng, G COD, D tên khách."
        )

    ref_col = col_map["reference_code"]
    cod_col = col_map.get("cod_amount", 6 if file_format == "shop_export" else 15)
    parsed: list[dict[str, Any]] = []

    data_rows = raw_rows[header_idx + 1 :]
    data_rows_alt = raw_rows_alt[header_idx + 1 :] if len(raw_rows_alt) > header_idx else []

    for offset, row in enumerate(data_rows):
        row_idx = header_idx + 2 + offset
        row_alt = data_rows_alt[offset] if offset < len(data_rows_alt) else None
        ref_code = _cell_str(row[ref_col] if ref_col < len(row) else "").upper()
        cod_raw = row[cod_col] if cod_col < len(row) else None
        cod_amount = _parse_cod_amount(cod_raw)

        if file_format == "shop_export":
            customer_name = _cell_str(row[col_map.get("customer_name", 3)] if col_map.get("customer_name", 3) < len(row) else "")
            phone = _cell_str(row[col_map.get("phone", 5)] if col_map.get("phone", 5) < len(row) else "")
            address = _cell_str(row[col_map.get("address", 4)] if col_map.get("address", 4) < len(row) else "")
            ma_sp_raw, product_code = read_ma_sp_from_excel_row(row, col_map, row_alt=row_alt)
            order_direct = row[col_map.get("order_code", 8)] if col_map.get("order_code", 8) < len(row) else None
            recipient_label = _build_recipient_label(
                customer_name=customer_name,
                phone=phone,
                address=address,
            )
            order_code = _resolve_order_code(
                direct=order_direct,
                product_code=ma_sp_raw or product_code,
                recipient_label=recipient_label,
            )
        else:
            ma_sp_raw, product_code = read_ma_sp_from_excel_row(row, col_map, row_alt=row_alt)
            recipient_col = col_map.get("recipient_label", 9)
            recipient_label = _cell_str(row[recipient_col] if recipient_col < len(row) else "")
            order_code = _resolve_order_code(
                product_code=ma_sp_raw or product_code,
                recipient_label=recipient_label,
            )

        if not ref_code and not recipient_label and not order_code:
            continue
        parsed.append(
            {
                "row_number": row_idx,
                "reference_code": ref_code,
                "recipient_label": recipient_label,
                "product_code": product_code or "",
                "product_code_raw": ma_sp_raw,
                "order_code": order_code,
                "cod_amount": cod_amount,
            }
        )

    if not parsed:
        if file_format == "shop_export":
            warnings.append(
                "Không có dòng dữ liệu hợp lệ (cột A mã vận đơn / cột I đơn hàng / cột G COD)."
            )
        else:
            warnings.append(
                "Không có dòng dữ liệu hợp lệ (cột MA_DON_HANG mã tham chiếu / TEN_NGUOI_NHAN)."
            )
    return parsed, warnings


def _ems_phase(description: Optional[str]) -> str:
    text = (description or "").lower()
    if "phát thành công" in text or "delivered successfully" in text:
        return "delivered"
    if "[cod]trả tiền" in text.replace(" ", ""):
        return "cod_settled"
    if "[cod]đã thu tiền" in text.replace(" ", ""):
        return "cod_collected"
    if "giao bưu tá" in text or "out for delivery" in text:
        return "out_for_delivery"
    if "vận chuyển" in text or "đến bưu cục" in text or "arrival at po" in text:
        return "in_transit"
    if "chấp nhận gửi" in text or "posting / collection" in text:
        return "posted"
    return "unknown"


def _parse_traced_at_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:10] if fmt == "%Y-%m-%d" else text, fmt).date()
        except ValueError:
            continue
    if len(text) >= 10 and text[4:5] == "-":
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            return None
    return None


def _merge_cod_paid_from_ems_tracking(result: dict[str, Any], ems: dict[str, Any]) -> None:
    """Không ghi cod_paid_* từ hành trình EMS — chỉ file đối soát COD quyết định EMS trả shop.

    [COD]Đã thu tiền / Đã thu tiền bưu tá = khách trả bưu tá, chưa phải EMS trả shop.
    """
    return


def _ems_only_status_message(ems_phase: str, ems_status: str) -> str:
    """Mô tả trạng thái EMS khi không có / không khớp đơn shop."""
    status = (ems_status or "").strip()
    if ems_phase in ("delivered", "cod_collected", "cod_settled"):
        return status or "EMS đã phát thành công."
    if ems_phase in ("out_for_delivery", "in_transit", "posted"):
        return status or "EMS đang vận chuyển."
    if status:
        return status
    return "EMS chưa có cập nhật hành trình rõ ràng."


EMS_SHOP_HANDOFF_MESSAGE = "Đã gửi EMS — đang giao tới bạn."


def _shop_ems_handoff_message(order_code: Optional[str] = None) -> str:
    code = (order_code or "").strip().upper()
    if code:
        return f"{code} · {EMS_SHOP_HANDOFF_MESSAGE}"
    return EMS_SHOP_HANDOFF_MESSAGE


def _try_sync_shop_on_ems_import(
    db: Session,
    order: Order,
    *,
    admin_id: Optional[int],
    reference_code: str,
    ems_tracking_code: Optional[str] = None,
    ems_status_description: Optional[str] = None,
    ems_phase: str = "posted",
) -> tuple[bool, str]:
    """Cập nhật đơn shop → đã gửi EMS giao nội địa khi import vào danh sách vận chuyển."""
    phase = (ems_phase or "posted").strip().lower()
    if phase not in shipment_svc.EMS_IMPORT_SYNC_PHASES:
        phase = "posted"
    tracking = (ems_tracking_code or reference_code or "").strip() or None
    synced, sync_msg = shipment_svc.apply_ems_import_shipping_sync(
        db,
        order,
        admin_id,
        ems_phase=phase,
        ems_tracking_code=tracking,
        ems_status_description=ems_status_description,
    )
    if synced:
        return True, _shop_ems_handoff_message(order.order_code)
    return False, sync_msg


def _order_not_found_message(*, order_code: Optional[str], ems_phase: str, ems_status: str) -> str:
    code = (order_code or "").strip().upper()
    ems_part = _ems_only_status_message(ems_phase, ems_status)
    if code and _is_shop_order_code(code):
        return (
            f"Đã lưu vận đơn — chưa có đơn shop {code} trên hệ thống. "
            f"Dùng cho đối soát COD/cước. {ems_part}"
        )
    if code and _DC_CODE_RE.fullmatch(code):
        return (
            f"Đã lưu vận đơn — mã tham chiếu {code} (Deal/CRM, không phải đơn 188.com.vn). "
            f"Dùng cho đối soát COD/cước. {ems_part}"
        )
    if code:
        return (
            f"Đã lưu vận đơn — mã cột I «{code}» không phải đơn shop hợp lệ. "
            f"Dùng cho đối soát COD/cước. {ems_part}"
        )
    return (
        f"Đã lưu vận đơn — thiếu mã đơn ở cột I (DHxxx/DCxxx). "
        f"Dùng cho đối soát COD/cước. {ems_part}"
    )


def _apply_unlinked_result(
    result: dict[str, Any],
    *,
    order_code: Optional[str],
    ems_phase: str = "unknown",
    ems_status: str = "",
    ems_error: Optional[str] = None,
) -> dict[str, Any]:
    """Import thành công nhưng chưa ghép được đơn shop — vẫn lưu COD/EMS để đối soát sau."""
    has_ems = ems_phase not in ("unknown", "") or bool((ems_status or "").strip())
    result["sync_status"] = "in_progress" if has_ems and not ems_error else "unlinked"
    result["sync_message"] = _order_not_found_message(
        order_code=order_code,
        ems_phase=ems_phase,
        ems_status=ems_status if not ems_error else "",
    )
    if ems_error:
        result["ems_error"] = ems_error
    return result


def _compare_ems_with_order(
    *,
    ems_phase: str,
    order_status: Optional[str],
    current_step_key: Optional[str],
) -> tuple[str, str]:
    st = (order_status or "").strip().lower()
    step = (current_step_key or "").strip().lower()

    if ems_phase in ("delivered", "cod_collected", "cod_settled"):
        if st in (OrderStatus.DELIVERED.value, OrderStatus.COMPLETED.value):
            return "matched", "EMS đã phát — đơn shop đã giao/hoàn tất."
        if st == OrderStatus.SHIPPING.value or step in ("domestic_shipping", "awaiting_confirm"):
            return "in_progress", "EMS đã phát — đơn shop đang giao/chờ khách xác nhận."
        return "mismatch", "EMS đã phát thành công nhưng đơn shop chưa cập nhật trạng thái giao."

    if ems_phase in ("out_for_delivery", "in_transit", "posted"):
        if st in (OrderStatus.SHIPPING.value, OrderStatus.DELIVERED.value, OrderStatus.COMPLETED.value):
            return "in_progress", "EMS đang vận chuyển/phát — đơn shop đã ở giai đoạn giao."
        if step in ("domestic_shipping", "awaiting_confirm"):
            return "in_progress", "EMS đang vận chuyển — timeline shop đã tới bước giao nội địa."
        return "mismatch", "EMS đang vận chuyển nhưng đơn shop chưa cập nhật giao hàng."

    if st in (OrderStatus.DELIVERED.value, OrderStatus.COMPLETED.value):
        return "mismatch", "Đơn shop đã giao nhưng EMS chưa báo phát thành công."

    return "in_progress", "Đã gửi EMS — đang giao tới bạn."


def _order_timeline_summary(db: Session, order: Order) -> tuple[Optional[str], Optional[str]]:
    """Đọc timeline hiện tại — không gọi advance_auto_milestones (tránh ghi DB khi import)."""
    st = getattr(order.status, "value", order.status)
    events = (
        db.query(OrderShipmentEvent)
        .filter(OrderShipmentEvent.order_id == order.id)
        .order_by(OrderShipmentEvent.sort_order.asc())
        .all()
    )
    current_key = None
    for ev in events:
        if ev.status == "active":
            current_key = ev.step_key
            break
    return current_key, st


def _dedupe_rows_by_reference(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    """Gộp dòng trùng mã tham chiếu trong cùng file — giữ dòng cuối."""
    warnings: list[str] = []
    by_ref: dict[str, dict[str, Any]] = {}
    no_ref: list[dict[str, Any]] = []
    dup_count = 0

    for row in rows:
        ref = (row.get("reference_code") or "").strip().upper()
        if not ref:
            no_ref.append(row)
            continue
        if ref in by_ref:
            dup_count += 1
        by_ref[ref] = {**row, "reference_code": ref}

    if dup_count:
        warnings.append(
            f"Có {dup_count} dòng trùng mã tham chiếu trong file — cập nhật theo dòng cuối cùng."
        )

    deduped = no_ref + sorted(by_ref.values(), key=lambda r: r.get("row_number", 0))
    return deduped, warnings


def _ems_lookup_candidates(reference_code: str, order: Optional[Order]) -> list[str]:
    candidates: list[str] = []
    ref = (reference_code or "").strip().upper()
    if ref:
        candidates.append(ref)
    if order:
        saved = (order.tracking_number or "").strip().upper()
        if saved and saved not in candidates:
            candidates.append(saved)
    return candidates


def _fetch_ems_with_fallback(codes: list[str]) -> dict[str, Any]:
    last: dict[str, Any] = {"available": False, "error": "Thiếu mã vận đơn EMS."}
    for code in codes:
        payload = ems_tracking_svc.fetch_ems_tracking(code)
        last = payload
        if payload.get("events") or not payload.get("error"):
            return payload
    return last


def process_ems_import_row(
    db: Session,
    row: dict[str, Any],
    *,
    admin_id: Optional[int] = None,
    skip_ems_tracking: bool = False,
) -> dict[str, Any]:
    reference_code = (row.get("reference_code") or "").strip().upper()
    order_code = (row.get("order_code") or "").strip().upper() or None

    result: dict[str, Any] = {
        **row,
        "reference_code": reference_code,
        "order_id": None,
        "order_status": None,
        "current_step_key": None,
        "tracking_number_saved": None,
        "ems_tracking_code": None,
        "ems_reference_code": None,
        "ems_status": None,
        "ems_phase": None,
        "sync_status": "pending",
        "sync_message": "",
        "ems_error": None,
        "order_synced": False,
        "order_sync_message": "",
        "cod_amount": row.get("cod_amount"),
    }

    if not reference_code:
        result["sync_status"] = "parse_error"
        result["sync_message"] = "Thiếu mã vận đơn (cột A / MA_DON_HANG)."
        return result

    order = (
        crud.order.get_order_by_code(db, order_code)
        if _is_shop_order_code(order_code)
        else None
    )

    if skip_ems_tracking:
        result["ems_reference_code"] = reference_code
        if order:
            result["order_id"] = order.id
            result["order_status"] = getattr(order.status, "value", order.status)
            result["tracking_number_saved"] = order.tracking_number
            step_key, _ = _order_timeline_summary(db, order)
            result["current_step_key"] = step_key
            synced, handoff_msg = _try_sync_shop_on_ems_import(
                db,
                order,
                admin_id=admin_id,
                reference_code=reference_code,
            )
            result["order_synced"] = synced
            result["order_sync_message"] = handoff_msg
            if synced:
                db.refresh(order)
                result["order_status"] = getattr(order.status, "value", order.status)
                result["tracking_number_saved"] = order.tracking_number
                step_key, _ = _order_timeline_summary(db, order)
                result["current_step_key"] = step_key
            result["sync_status"] = "in_progress"
            result["sync_message"] = handoff_msg
            return result
        return _apply_unlinked_result(
            result,
            order_code=order_code,
            ems_phase="unknown",
            ems_status="",
        )

    ems = _fetch_ems_with_fallback(_ems_lookup_candidates(reference_code, order))
    has_ems = not (ems.get("error") and not ems.get("events"))

    if has_ems:
        ems_status = ems.get("current_status_description") or ""
        ems_phase = _ems_phase(ems_status)
        result["ems_tracking_code"] = ems.get("tracking_code")
        result["ems_reference_code"] = ems.get("reference_code") or reference_code
        result["ems_status"] = ems_status
        result["ems_phase"] = ems_phase
        _merge_cod_paid_from_ems_tracking(result, ems)
    else:
        ems_status = ""
        ems_phase = "unknown"
        result["ems_error"] = ems.get("error")
        result["ems_reference_code"] = reference_code

    if not order:
        return _apply_unlinked_result(
            result,
            order_code=order_code,
            ems_phase=ems_phase,
            ems_status=ems_status,
            ems_error=result.get("ems_error"),
        )

    result["order_id"] = order.id
    result["order_status"] = getattr(order.status, "value", order.status)
    result["tracking_number_saved"] = order.tracking_number
    step_key, _ = _order_timeline_summary(db, order)
    result["current_step_key"] = step_key

    if not has_ems:
        synced, handoff_msg = _try_sync_shop_on_ems_import(
            db,
            order,
            admin_id=admin_id,
            reference_code=reference_code,
        )
        result["order_synced"] = synced
        result["order_sync_message"] = handoff_msg
        if synced:
            db.refresh(order)
            result["order_status"] = getattr(order.status, "value", order.status)
            result["tracking_number_saved"] = order.tracking_number
            step_key, _ = _order_timeline_summary(db, order)
            result["current_step_key"] = step_key
        result["sync_status"] = "in_progress"
        result["sync_message"] = handoff_msg
        return result

    if ems_phase != "unknown":
        synced, sync_msg = _try_sync_shop_on_ems_import(
            db,
            order,
            admin_id=admin_id,
            reference_code=reference_code,
            ems_tracking_code=result.get("ems_tracking_code"),
            ems_status_description=ems_status,
            ems_phase=ems_phase,
        )
        result["order_synced"] = synced
        result["order_sync_message"] = sync_msg
        if synced:
            db.refresh(order)
            result["order_status"] = getattr(order.status, "value", order.status)
            result["tracking_number_saved"] = order.tracking_number
            step_key, _ = _order_timeline_summary(db, order)
            result["current_step_key"] = step_key

    sync_status, sync_message = _compare_ems_with_order(
        ems_phase=ems_phase,
        order_status=result["order_status"],
        current_step_key=result["current_step_key"],
    )
    result["sync_status"] = sync_status
    result["sync_message"] = (
        result["order_sync_message"]
        if result.get("order_synced") and result.get("order_sync_message")
        else sync_message
    )
    if result.get("order_sync_message") and not result.get("order_synced"):
        result["sync_message"] = result["order_sync_message"]

    saved_tracking = (order.tracking_number or "").strip().upper()
    ems_tracking = (result["ems_tracking_code"] or "").strip().upper()
    if ems_tracking and saved_tracking and saved_tracking != ems_tracking:
        result["sync_status"] = "mismatch"
        result["sync_message"] = (
            f"Mã vận đơn shop ({saved_tracking}) khác EMS ({ems_tracking}). "
            + result["sync_message"]
        )

    return result


_SYNC_STATUS_ORDER = (
    "matched",
    "in_progress",
    "mismatch",
    "unlinked",
    "order_not_found",
    "ems_not_found",
    "parse_error",
    "pending",
)


def _build_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    counts: dict[str, int] = {k: 0 for k in _SYNC_STATUS_ORDER}
    cod_totals: dict[str, int] = {k: 0 for k in _SYNC_STATUS_ORDER}

    for row in rows:
        status = (row.get("sync_status") or "pending").strip()
        if status not in counts:
            counts[status] = 0
            cod_totals[status] = 0
        counts[status] += 1
        cod_raw = row.get("cod_amount")
        if cod_raw is not None:
            try:
                cod_totals[status] += int(cod_raw)
            except (TypeError, ValueError):
                pass

    breakdown: list[dict[str, Any]] = [
        {"key": key, "count": counts.get(key, 0), "cod_total": cod_totals.get(key, 0)}
        for key in _SYNC_STATUS_ORDER
        if counts.get(key, 0) > 0
    ]
    for key, count in counts.items():
        if count > 0 and key not in _SYNC_STATUS_ORDER:
            breakdown.append({"key": key, "count": count, "cod_total": cod_totals.get(key, 0)})

    total_cod = sum(cod_totals.values())

    unlinked_count = counts.get("unlinked", 0) + counts.get("order_not_found", 0)

    return {
        "total_rows": len(rows),
        "matched": counts.get("matched", 0),
        "in_progress": counts.get("in_progress", 0),
        "mismatch": counts.get("mismatch", 0),
        "unlinked": unlinked_count,
        "order_not_found": counts.get("order_not_found", 0),
        "ems_not_found": counts.get("ems_not_found", 0),
        "parse_error": counts.get("parse_error", 0),
        "total_cod_amount": total_cod,
        "breakdown": breakdown,
    }


def sanitize_cod_paid_without_settlement(db: Session) -> int:
    """Xóa cod_paid_* ghi nhầm từ EMS tracking — chỉ giữ sau import file đối soát COD."""
    rows = (
        db.query(EmsShippingRecord)
        .filter(
            (EmsShippingRecord.cod_paid_amount.isnot(None)) | (EmsShippingRecord.cod_paid_date.isnot(None))
        )
        .all()
    )
    fixed = 0
    for record in rows:
        if (record.cod_settlement_status or "").strip().lower() == "matched":
            continue
        record.cod_paid_amount = None
        record.cod_paid_date = None
        fixed += 1
    if fixed:
        db.commit()
    return fixed


def sanitize_invalid_ems_product_codes(db: Session) -> int:
    """Xóa product_code ghi nhầm từ nhãn người nhận — chỉ giữ MA_SP cột H hợp lệ."""
    rows = db.query(EmsShippingRecord).filter(EmsShippingRecord.product_code.isnot(None)).all()
    fixed = 0
    for record in rows:
        raw = (record.product_code or "").strip()
        if not raw:
            continue
        normalized = warehouse_sku_from_col_h_cell(raw)
        if normalized != raw or looks_like_recipient_not_sku(raw):
            record.product_code = normalized or None
            fixed += 1
    if fixed:
        db.commit()
    return fixed


def _record_to_dict(record: EmsShippingRecord) -> dict[str, Any]:
    settlement = (record.cod_settlement_status or "").strip().lower()
    cod_paid_amount = int(record.cod_paid_amount) if record.cod_paid_amount is not None else None
    cod_paid_date = _isoformat_value(record.cod_paid_date)
    if settlement != "matched":
        cod_paid_amount = None
        cod_paid_date = None
    return _apply_return_to_shop_fields(
        {
            "id": record.id,
            "row_number": record.excel_row_number or 0,
            "reference_code": record.reference_code or "",
            "recipient_label": record.recipient_label or "",
            "order_code": record.order_code,
            "order_id": record.order_id,
            "order_status": record.order_status,
            "current_step_key": record.current_step_key,
            "tracking_number_saved": record.tracking_number_saved,
            "ems_tracking_code": record.ems_tracking_code,
            "ems_reference_code": record.ems_reference_code,
            "ems_status": record.ems_status,
            "ems_phase": record.ems_phase,
            "sync_status": record.sync_status,
            "sync_message": record.sync_message or "",
            "ems_error": record.ems_error,
            "cod_amount": int(record.cod_amount) if record.cod_amount is not None else None,
            "cod_paid_amount": cod_paid_amount,
            "cod_paid_date": cod_paid_date,
            "cod_settlement_status": record.cod_settlement_status,
            "cod_settlement_message": record.cod_settlement_message,
            "freight_amount": int(record.freight_amount) if record.freight_amount is not None else None,
            "freight_settled_at": _isoformat_value(record.freight_settled_at),
            "freight_settlement_status": record.freight_settlement_status,
            "freight_settlement_message": record.freight_settlement_message,
            "freight_high_fee_warning": record.freight_high_fee_warning,
            "return_to_shop_label": None,
            "shop_return_received_at": _isoformat_value(record.shop_return_received_at),
            "order_synced": False,
            "order_sync_message": "",
        }
    )


def _result_has_ems_tracking_payload(result: dict[str, Any]) -> bool:
    """True khi đã tra EMS thật (không phải import nhanh chỉ lưu Excel)."""
    return bool(
        (result.get("ems_tracking_code") or "").strip()
        or (result.get("ems_status") or "").strip()
        or (result.get("ems_phase") or "").strip()
    )


def _upsert_record(
    db: Session,
    result: dict[str, Any],
    *,
    admin_id: Optional[int] = None,
    source_filename: Optional[str] = None,
) -> tuple[EmsShippingRecord, bool]:
    ref = (result.get("reference_code") or "").strip().upper()
    if not ref:
        raise ValueError("Thiếu mã tham chiếu để lưu bảng vận chuyển.")

    record = db.query(EmsShippingRecord).filter(EmsShippingRecord.reference_code == ref).first()
    created = record is None
    if not record:
        record = EmsShippingRecord(reference_code=ref)
        db.add(record)

    record.recipient_label = result.get("recipient_label") or ""
    pc = warehouse_sku_from_col_h_cell(result.get("product_code")) or ""
    if not pc:
        pc = warehouse_sku_from_col_h_cell(result.get("product_code_raw")) or ""
    if pc:
        record.product_code = pc
    elif created:
        record.product_code = None
    # Cập nhật import: không xóa SKU cũ nếu dòng mới không đọc được cột H
    record.order_code = result.get("order_code")
    record.order_id = result.get("order_id")
    record.excel_row_number = result.get("row_number")
    # Giữ xác nhận hoàn shop — cron/import EMS không được ghi đè về None/shipping.
    shop_confirmed = getattr(record, "shop_return_received_at", None) is not None or (
        (record.order_status or "").strip().lower() == OrderStatus.RETURNED.value
    )
    if shop_confirmed:
        record.order_status = OrderStatus.RETURNED.value
        if getattr(record, "shop_return_received_at", None) is None:
            record.shop_return_received_at = record.updated_at or datetime.now(timezone.utc)
    elif result.get("order_status") is not None:
        record.order_status = result.get("order_status")
    record.current_step_key = result.get("current_step_key")
    record.tracking_number_saved = result.get("tracking_number_saved")

    has_ems = _result_has_ems_tracking_payload(result)
    if created:
        record.ems_tracking_code = result.get("ems_tracking_code")
        record.ems_status = result.get("ems_status")
        record.ems_phase = result.get("ems_phase")
        record.ems_error = result.get("ems_error")
        record.sync_status = result.get("sync_status") or "pending"
        record.sync_message = result.get("sync_message")
    elif has_ems:
        tracking = (result.get("ems_tracking_code") or "").strip()
        if tracking:
            record.ems_tracking_code = tracking
        if result.get("ems_status") is not None:
            record.ems_status = result.get("ems_status")
        if result.get("ems_phase") is not None:
            record.ems_phase = result.get("ems_phase")
        record.ems_error = result.get("ems_error")
        if result.get("sync_status"):
            record.sync_status = result.get("sync_status") or "pending"
        if not shop_confirmed and result.get("sync_message") is not None:
            record.sync_message = result.get("sync_message")
    elif not (record.sync_status or "").strip():
        record.sync_status = result.get("sync_status") or "pending"
        record.sync_message = result.get("sync_message")

    ems_ref = (result.get("ems_reference_code") or ref or "").strip() or None
    if ems_ref and (created or has_ems or not (record.ems_reference_code or "").strip()):
        record.ems_reference_code = ems_ref
    cod = result.get("cod_amount")
    record.cod_amount = int(cod) if cod is not None else None
    settlement = (record.cod_settlement_status or "").strip().lower()
    # COD EMS trả shop — chỉ từ file đối soát (cod_settlement_status=matched).
    if settlement != "matched":
        record.cod_paid_amount = None
        record.cod_paid_date = None
    if source_filename:
        record.import_source_filename = source_filename
    if admin_id:
        record.imported_by_admin_id = admin_id
    _apply_return_to_shop_on_record(record)
    db.flush()
    return record, created


def _apply_return_to_shop_fields(row: dict[str, Any]) -> dict[str, Any]:
    from app.services.shipping_operations import (
        RETURN_PENDING_SHOP_LABEL,
        RETURN_SHOP_RECEIVED_LABEL,
        return_to_shop_label,
    )

    label = return_to_shop_label(
        order_status=row.get("order_status"),
        ems_status=row.get("ems_status"),
        shop_return_received_at=row.get("shop_return_received_at"),
    )
    row["return_to_shop_label"] = label
    if label == RETURN_PENDING_SHOP_LABEL:
        row["sync_message"] = RETURN_PENDING_SHOP_LABEL
    elif label == RETURN_SHOP_RECEIVED_LABEL:
        row["sync_message"] = RETURN_SHOP_RECEIVED_LABEL
    return row


def _apply_return_to_shop_on_record(record: EmsShippingRecord) -> None:
    from app.services.shipping_operations import (
        RETURN_PENDING_SHOP_LABEL,
        RETURN_SHOP_RECEIVED_LABEL,
        return_to_shop_label,
    )

    label = return_to_shop_label(
        order_status=record.order_status,
        ems_status=record.ems_status,
        shop_return_received_at=getattr(record, "shop_return_received_at", None),
    )
    if label == RETURN_PENDING_SHOP_LABEL:
        record.sync_message = RETURN_PENDING_SHOP_LABEL
    elif label == RETURN_SHOP_RECEIVED_LABEL:
        record.sync_message = RETURN_SHOP_RECEIVED_LABEL


def _enrich_row_from_live_order(db: Session, row: dict[str, Any]) -> dict[str, Any]:
    """Cập nhật trạng thái đơn shop mới nhất khi hiển thị bảng (không cần import lại)."""
    if row.get("shop_return_received_at") or (
        (row.get("order_status") or "").strip().lower() == OrderStatus.RETURNED.value
    ):
        row["order_status"] = OrderStatus.RETURNED.value
        return _apply_return_to_shop_fields(row)

    order_code = (row.get("order_code") or "").strip().upper()
    if not _is_shop_order_code(order_code):
        return _apply_return_to_shop_fields(row)
    order = crud.order.get_order_by_code(db, order_code)
    if not order:
        return _apply_return_to_shop_fields(row)
    row["order_id"] = order.id
    st = getattr(order.status, "value", order.status)
    if (st or "").strip().lower() == OrderStatus.RETURNED.value:
        row["order_status"] = OrderStatus.RETURNED.value
        return _apply_return_to_shop_fields(row)
    row["order_status"] = st
    row["tracking_number_saved"] = order.tracking_number
    step_key, _ = _order_timeline_summary(db, order)
    row["current_step_key"] = step_key
    ems_phase = (row.get("ems_phase") or "").strip()
    if ems_phase and ems_phase != "unknown":
        sync_status, sync_message = _compare_ems_with_order(
            ems_phase=ems_phase,
            order_status=row["order_status"],
            current_step_key=step_key,
        )
        row["sync_status"] = sync_status
        row["sync_message"] = sync_message
    elif row.get("order_id") and (row.get("reference_code") or row.get("ems_reference_code")):
        row["sync_status"] = "in_progress"
        row["sync_message"] = _shop_ems_handoff_message(order_code)
    return _apply_return_to_shop_fields(row)


def _build_summary_from_db(db: Session) -> dict[str, Any]:
    rows = (
        db.query(
            EmsShippingRecord.sync_status,
            func.count(EmsShippingRecord.id),
            func.coalesce(func.sum(EmsShippingRecord.cod_amount), 0),
        )
        .group_by(EmsShippingRecord.sync_status)
        .all()
    )
    counts: dict[str, int] = {k: 0 for k in _SYNC_STATUS_ORDER}
    cod_totals: dict[str, int] = {k: 0 for k in _SYNC_STATUS_ORDER}
    for status_raw, count_raw, cod_raw in rows:
        status = (status_raw or "pending").strip()
        if status not in counts:
            counts[status] = 0
            cod_totals[status] = 0
        counts[status] += int(count_raw or 0)
        cod_totals[status] += int(cod_raw or 0)

    breakdown: list[dict[str, Any]] = [
        {"key": key, "count": counts.get(key, 0), "cod_total": cod_totals.get(key, 0)}
        for key in _SYNC_STATUS_ORDER
        if counts.get(key, 0) > 0
    ]
    for key, count in counts.items():
        if count > 0 and key not in _SYNC_STATUS_ORDER:
            breakdown.append({"key": key, "count": count, "cod_total": cod_totals.get(key, 0)})

    unlinked_count = counts.get("unlinked", 0) + counts.get("order_not_found", 0)
    return {
        "total_rows": sum(counts.values()),
        "matched": counts.get("matched", 0),
        "in_progress": counts.get("in_progress", 0),
        "mismatch": counts.get("mismatch", 0),
        "unlinked": unlinked_count,
        "order_not_found": counts.get("order_not_found", 0),
        "ems_not_found": counts.get("ems_not_found", 0),
        "parse_error": counts.get("parse_error", 0),
        "total_cod_amount": sum(cod_totals.values()),
        "breakdown": breakdown,
    }


def _apply_search_filter(query, search: Optional[str]):
    term = (search or "").strip()
    if not term:
        return query
    like = f"%{term.upper()}%"
    return query.filter(
        or_(
            func.upper(EmsShippingRecord.reference_code).like(like),
            func.upper(EmsShippingRecord.order_code).like(like),
            func.upper(EmsShippingRecord.ems_tracking_code).like(like),
            func.upper(EmsShippingRecord.ems_reference_code).like(like),
            func.upper(EmsShippingRecord.tracking_number_saved).like(like),
        )
    )


def _apply_sync_status_filter(query, sync_status: Optional[str]):
    status = (sync_status or "").strip()
    if not status or status == "all":
        return query
    if status == "unlinked":
        return query.filter(
            or_(
                EmsShippingRecord.sync_status == "unlinked",
                EmsShippingRecord.sync_status == "order_not_found",
            )
        )
    return query.filter(EmsShippingRecord.sync_status == status)


_EMS_REFRESH_MAX_IDS = 500
_EMS_TERMINAL_PHASES = frozenset({"delivered", "cod_collected", "cod_settled"})


def _apply_non_terminal_refresh_filter(query):
    """Chỉ tra lại đơn chưa ở trạng thái EMS cuối (đã giao / thu COD xong)."""
    terminal = tuple(_EMS_TERMINAL_PHASES)
    return query.filter(
        or_(
            EmsShippingRecord.ems_phase.is_(None),
            EmsShippingRecord.ems_phase == "",
            func.lower(EmsShippingRecord.ems_phase) == "unknown",
            ~func.lower(EmsShippingRecord.ems_phase).in_(terminal),
            and_(
                EmsShippingRecord.ems_error.isnot(None),
                EmsShippingRecord.ems_error != "",
            ),
        )
    )


def collect_record_ids_for_refresh(
    db: Session,
    *,
    search: Optional[str] = None,
    sync_status: Optional[str] = None,
    non_terminal_only: bool = False,
    limit: int = _EMS_REFRESH_MAX_IDS,
) -> list[int]:
    """Lấy id các dòng EMS cần tra lại (theo tìm kiếm / bộ lọc)."""
    query = db.query(EmsShippingRecord.id)
    status = (sync_status or "").strip()
    if status and status != "all":
        query = _apply_sync_status_filter(query, status)
    search_term = (search or "").strip() or None
    query = _apply_search_filter(query, search_term)
    if non_terminal_only:
        query = _apply_non_terminal_refresh_filter(query)
    cap = max(1, min(int(limit or _EMS_REFRESH_MAX_IDS), _EMS_REFRESH_MAX_IDS))
    rows = (
        query.order_by(
            EmsShippingRecord.updated_at.asc(),
            EmsShippingRecord.id.asc(),
        )
        .limit(cap)
        .all()
    )
    return [int(row[0]) for row in rows]


def list_ems_shipping_records(
    db: Session,
    *,
    skip: int = 0,
    limit: int = 50,
    sync_status: Optional[str] = None,
    search: Optional[str] = None,
) -> dict[str, Any]:
    skip = max(0, int(skip or 0))
    limit = max(1, min(int(limit or 50), 200))
    search_term = (search or "").strip() or None

    base_query = db.query(EmsShippingRecord)
    filtered_query = _apply_sync_status_filter(base_query, sync_status)
    filtered_query = _apply_search_filter(filtered_query, search_term)
    total = int(base_query.count() or 0)
    filtered_total = int(filtered_query.count() or 0)

    records = (
        filtered_query.order_by(
            EmsShippingRecord.updated_at.desc(),
            EmsShippingRecord.id.desc(),
        )
        .offset(skip)
        .limit(limit)
        .all()
    )
    rows = [_enrich_row_from_live_order(db, _record_to_dict(r)) for r in records]
    return {
        "ok": True,
        "warnings": [],
        "summary": _build_summary_from_db(db),
        "pagination": {
            "skip": skip,
            "limit": limit,
            "total": total,
            "filtered_total": filtered_total,
            "search": search_term,
        },
        "rows": rows,
    }


def _import_batch_row_to_report_row(row: EmsShippingImportBatchRow) -> dict[str, Any]:
    return {
        "id": row.ems_shipping_record_id,
        "row_number": row.excel_row_number or 0,
        "reference_code": row.reference_code or "",
        "recipient_label": row.recipient_label or "",
        "order_code": row.order_code,
        "order_id": row.order_id,
        "order_status": None,
        "current_step_key": None,
        "tracking_number_saved": None,
        "ems_tracking_code": None,
        "ems_reference_code": None,
        "ems_status": None,
        "ems_phase": None,
        "sync_status": row.sync_status or "pending",
        "sync_message": row.sync_message or "",
        "ems_error": None,
        "cod_amount": int(row.cod_amount) if row.cod_amount is not None else None,
        "import_action": row.import_action,
    }


def _batch_to_dict(batch: EmsShippingImportBatch, report_rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "id": batch.id,
        "source_filename": batch.source_filename,
        "created_at": _isoformat_value(batch.created_at),
        "file_rows_processed": int(batch.file_rows_processed or 0),
        "order_count": int(batch.order_count or 0),
        "created_count": int(batch.created_count or 0),
        "updated_count": int(batch.updated_count or 0),
        "skipped_no_reference_count": int(batch.skipped_no_reference_count or 0),
        "orders_synced_count": int(batch.orders_synced_count or 0),
        "total_cod_amount": int(batch.total_cod_amount or 0),
        "import_report": {
            "order_count": int(batch.order_count or 0),
            "total_cod_amount": int(batch.total_cod_amount or 0),
            "created": int(batch.created_count or 0),
            "updated": int(batch.updated_count or 0),
            "skipped_no_reference": int(batch.skipped_no_reference_count or 0),
            "orders_synced": int(batch.orders_synced_count or 0),
            "rows": report_rows,
        },
    }


def _persist_import_batch(
    db: Session,
    *,
    import_report_rows: list[dict[str, Any]],
    file_rows_processed: int,
    created: int,
    updated: int,
    skipped_no_reference: int,
    orders_synced: int,
    admin_id: Optional[int] = None,
    source_filename: Optional[str] = None,
) -> EmsShippingImportBatch:
    total_cod = sum(int(r.get("cod_amount") or 0) for r in import_report_rows)
    batch = EmsShippingImportBatch(
        source_filename=source_filename,
        imported_by_admin_id=admin_id,
        file_rows_processed=file_rows_processed,
        order_count=len(import_report_rows),
        created_count=created,
        updated_count=updated,
        skipped_no_reference_count=skipped_no_reference,
        orders_synced_count=orders_synced,
        total_cod_amount=total_cod,
    )
    db.add(batch)
    db.flush()
    for row_dict in import_report_rows:
        db.add(
            EmsShippingImportBatchRow(
                batch_id=batch.id,
                ems_shipping_record_id=row_dict.get("id"),
                excel_row_number=row_dict.get("row_number"),
                reference_code=(row_dict.get("reference_code") or "").strip() or None,
                recipient_label=row_dict.get("recipient_label"),
                order_code=row_dict.get("order_code"),
                order_id=row_dict.get("order_id"),
                cod_amount=row_dict.get("cod_amount"),
                import_action=row_dict.get("import_action"),
                sync_status=row_dict.get("sync_status"),
                sync_message=row_dict.get("sync_message"),
            )
        )
    return batch


def list_ems_import_batches(db: Session, *, limit: int = 30) -> dict[str, Any]:
    limit = max(1, min(int(limit or 30), 100))
    batches = (
        db.query(EmsShippingImportBatch)
        .order_by(EmsShippingImportBatch.created_at.desc(), EmsShippingImportBatch.id.desc())
        .limit(limit)
        .all()
    )
    items: list[dict[str, Any]] = []
    for batch in batches:
        rows = (
            db.query(EmsShippingImportBatchRow)
            .filter(EmsShippingImportBatchRow.batch_id == batch.id)
            .order_by(EmsShippingImportBatchRow.excel_row_number.asc(), EmsShippingImportBatchRow.id.asc())
            .all()
        )
        report_rows = [_import_batch_row_to_report_row(r) for r in rows]
        items.append(_batch_to_dict(batch, report_rows))
    return {
        "ok": True,
        "batches": items,
        "import_batch": items[0] if items else None,
    }


def delete_ems_shipping_records(db: Session, record_ids: list[int]) -> int:
    ids = [int(x) for x in record_ids if int(x) > 0]
    if not ids:
        return 0
    deleted = (
        db.query(EmsShippingRecord)
        .filter(EmsShippingRecord.id.in_(ids))
        .delete(synchronize_session=False)
    )
    db.commit()
    return int(deleted or 0)


def import_ems_shipment_excel(
    db: Session,
    file_bytes: bytes,
    *,
    admin_id: Optional[int] = None,
    source_filename: Optional[str] = None,
) -> dict[str, Any]:
    rows, warnings = parse_ems_export_rows(file_bytes)
    rows, dedupe_warnings = _dedupe_rows_by_reference(rows)
    warnings.extend(dedupe_warnings)

    missing_ma_sp = sum(1 for r in rows if not (r.get("product_code") or "").strip())
    if missing_ma_sp:
        warnings.append(
            f"{missing_ma_sp} dòng không đọc được MA_SP (cột H) — "
            "kiểm tra ô H / TEN_SP có dạng B7796/41/2-...; nếu ô H là công thức, mở file trong Excel rồi Save trước khi import."
        )

    # Luôn import nhanh — tra EMS chạy nền qua ems_tracking_refresh worker.
    warnings.append(
        "Import nhanh: đã lưu mã vận đơn/COD/đơn. Tra EMS sẽ chạy nền trên server theo thứ tự."
    )

    results = [
        process_ems_import_row(db, row, admin_id=admin_id, skip_ems_tracking=True)
        for row in rows
    ]

    created = updated = skipped_no_reference = orders_synced = 0
    imported_record_ids: list[int] = []
    import_report_rows: list[dict[str, Any]] = []
    for result in results:
        ref = (result.get("reference_code") or "").strip().upper()
        if not ref:
            skipped_no_reference += 1
            continue
        if result.get("order_synced"):
            orders_synced += 1
        record, was_created = _upsert_record(
            db,
            result,
            admin_id=admin_id,
            source_filename=source_filename,
        )
        imported_record_ids.append(int(record.id))
        if was_created:
            created += 1
        else:
            updated += 1
        row_dict = _enrich_row_from_live_order(db, _record_to_dict(record))
        row_dict["import_action"] = "created" if was_created else "updated"
        import_report_rows.append(row_dict)

    import_batch = _persist_import_batch(
        db,
        import_report_rows=import_report_rows,
        file_rows_processed=len(results),
        created=created,
        updated=updated,
        skipped_no_reference=skipped_no_reference,
        orders_synced=orders_synced,
        admin_id=admin_id,
        source_filename=source_filename,
    )

    db.commit()

    tracking_refresh_job_id = None
    try:
        from app.services import ems_tracking_refresh as ems_refresh_svc

        tracking_refresh_job_id = ems_refresh_svc.enqueue_tracking_refresh(
            imported_record_ids,
            admin_id=admin_id,
            source="import",
        )
    except Exception as exc:
        warnings.append(f"Không khởi chạy job tra EMS nền: {exc}")

    payload = list_ems_shipping_records(db, skip=0, limit=50)
    payload["warnings"] = warnings
    payload["import_stats"] = {
        "file_rows_processed": len(results),
        "created": created,
        "updated": updated,
        "skipped_no_reference": skipped_no_reference,
        "orders_synced": orders_synced,
    }
    batch_saved = _batch_to_dict(import_batch, import_report_rows)
    batches_payload = list_ems_import_batches(db)
    payload["batches"] = batches_payload["batches"]
    payload["import_batch"] = batch_saved
    payload["import_report"] = batch_saved["import_report"]
    payload["tracking_refresh_job_id"] = tracking_refresh_job_id
    return payload
