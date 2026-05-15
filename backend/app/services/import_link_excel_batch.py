"""

Đọc Excel batch **chỉ** mẫu export listing tái nhập (vd. `222222222222222.xlsx`): hai dòng nhãn đầu (EN/VI).



**Bắt buộc** tiêu đề có **Link** (`Link SP`, `item_url`…) và **Giá Tệ / China price**. Không hỗ trợ cột giá VND Q,

cột G AK, khối L–O hay layout Excel đặt hàng cũ.



- **shop_name_chinese**, **chinese_name**: nhận qua tiêu đề (`shop_name_chinese`, Shop Trung Quốc, …).

- **Giá Tệ** → `pro_lower_price` và **`price`** (VNĐ) = làm_tròn(CN¥ × hệ_số_IF × tỷ_giá), sau đó làm tròn **lên** bội 10.000 ₫ (`listing_cny_grid`).

  Tỷ giá nền `LISTING_IMPORT_VND_PER_CNY`; cột `vnd_per_cny_used` / Tỷ giá ghi đè theo dòng nếu có.

- **Mã sp** (SKU `[A-Z][0-9]{4}`): chỉ khi có cột tiêu đề «Mã sp»… — không đọc cột B cố định.



`merge_import_excel_overlay_into_product_data` vẫn hỗ trợ các khóa legacy nếu overlay cũ còn trong DB (không sinh từ parser này).

"""

from __future__ import annotations



import math

import re

import unicodedata

from pathlib import Path

from typing import Any, Dict, List, Optional, Tuple



from openpyxl import load_workbook



from app.services.listing_cny_grid import (

    DEFAULT_VND_PER_CNY_FOR_LISTING_ESTIMATE,

    cny_exchange_multiplier_from_grid,

    estimate_listing_vnd_rounded,

    parse_approx_cny_amount_from_cell,

)

from app.services.product_internal_sku import INTERNAL_SKU_RE



_ws_re = re.compile(r"\s+")



MAX_IMPORT_COL_1BASED = 32

DATA_FIRST_ROW = 2

_MAX_ROWS = 500





def _merged_labels_for_cols(

    header_row1: Tuple[Any, ...], header_row2: Tuple[Any, ...], max_col: int

) -> Dict[int, set[str]]:

    by_col: Dict[int, set[str]] = {}

    for j in range(1, max(1, max_col) + 1):

        lab = _labels_at(header_row1, header_row2, j)

        if lab:

            by_col[j] = lab

    return by_col





def _sorted_cols_matching(by_col: Dict[int, set[str]], needles: frozenset[str]) -> List[int]:

    return sorted([j for j, labs in by_col.items() if labs & needles])





def _first_col_matching(by_col: Dict[int, set[str]], needles: frozenset[str]) -> Optional[int]:

    xs = _sorted_cols_matching(by_col, needles)

    return xs[0] if xs else None





def _resolve_link_column_required(by_col: Dict[int, set[str]]) -> Optional[int]:

    tier_strong = frozenset(

        _norm_header(x)

        for x in (

            "Link SP",

            "link_sp",

            "Link sp",

            "item_url",

            "product_url",

            "product url",

            "URL sp",

            "url_san_pham",

        )

    )

    tier_weak = frozenset({_norm_header("link")})

    cols_s = _sorted_cols_matching(by_col, tier_strong)

    if cols_s:

        return min(cols_s)

    cols_w = _sorted_cols_matching(by_col, tier_weak)

    if len(cols_w) == 1:

        return cols_w[0]

    if len(cols_w) > 1:

        return min(cols_w)

    return None





def _resolve_china_price_column_required(by_col: Dict[int, set[str]]) -> Optional[int]:

    needles = frozenset(

        _norm_header(x)

        for x in (

            "China price",

            "china price",

            "Giá Tệ",

            "giá tệ",

            "gia te",

            "price_cny",

            "gia_te",

            "price cny",

        )

    )

    return _first_col_matching(by_col, needles)





def _resolve_optional_vnd_per_cny_column(by_col: Dict[int, set[str]]) -> Optional[int]:

    needles = frozenset(

        _norm_header(x)

        for x in ("vnd_per_cny_used", "vnd_per_cny", "ty gia", "tỷ giá", "ty_gia", "tygia")

    )

    return _first_col_matching(by_col, needles)





def _default_listing_import_vnd_per_cny() -> float:

    try:

        from app.core.config import settings



        r = float(getattr(settings, "LISTING_IMPORT_VND_PER_CNY", DEFAULT_VND_PER_CNY_FOR_LISTING_ESTIMATE))

        if math.isfinite(r) and r > 0:

            return r

    except Exception:

        pass

    return float(DEFAULT_VND_PER_CNY_FOR_LISTING_ESTIMATE)





def _resolve_shop_chinese_column(by_col: Dict[int, set[str]]) -> Optional[int]:

    needles = frozenset(

        _norm_header(x)

        for x in ("shop_name_chinese", "shop trung quốc", "shop trung quoc", "shop_cn", "supplier_cn")

    )

    return _first_col_matching(by_col, needles)





def _resolve_chinese_name_column(by_col: Dict[int, set[str]]) -> Optional[int]:

    needles = frozenset(

        _norm_header(x) for x in ("chinese_name", "ten tiếng trung", "tên tiếng trung", "tên tg", "tieude_goc")

    )

    return _first_col_matching(by_col, needles)





def _resolve_internal_sku_header_column(by_col: Dict[int, set[str]]) -> Optional[int]:

    sku_needles = frozenset(

        _norm_header(x) for x in ("Mã sp", "Ma sp", "mã sp", "internal sku", "mã sku", "ma sku", "SKU nội bộ")

    )

    sku_cols = _sorted_cols_matching(by_col, sku_needles)

    if len(sku_cols) == 1:

        return sku_cols[0]

    return None





def _row_has_http_link(url_raw: str) -> bool:

    u = url_raw.strip()

    return bool(u.startswith("http://") or u.startswith("https://"))





def _norm_header(text: Any) -> str:

    if text is None:

        return ""

    s = unicodedata.normalize("NFKC", str(text)).strip().casefold()

    return _ws_re.sub(" ", s)





def _cell_str(val: Any) -> str:

    if val is None:

        return ""

    if isinstance(val, float) and val.is_integer():

        return str(int(val))

    return str(val).strip()





_VND_CELL_NOISE_RE = re.compile(

    r"[\u00a0\s]|₫|\bVNĐ\b|\bvnd\b|đồng",

    re.IGNORECASE,

)





def _parse_vnd_amount_from_excel_cell(val: Any) -> Optional[float]:

    """Parse ô VND (legacy overlay `price_display_vnd_g`)."""

    if val is None:

        return None

    if isinstance(val, bool):

        return None

    if isinstance(val, int):

        return float(val)

    if isinstance(val, float):

        return float(val)

    s = _VND_CELL_NOISE_RE.sub("", str(val).strip()).replace("\u2212", "-")

    s = re.sub(r"(?i)[\s]*(?:đ|vnđ|[₫])[\s]*$", "", s).strip()

    if not s:

        return None

    if s.startswith("(") and s.endswith(")"):

        s = "-" + s[1:-1]

    comma_to_dot = s.replace(",", ".")

    dot_count = comma_to_dot.count(".")

    if dot_count >= 2:

        try:

            return float(comma_to_dot.replace(".", ""))

        except ValueError:

            return None

    if dot_count == 1:

        left, right = comma_to_dot.split(".", 1)

        if left.lstrip("-").isdigit() and right.isdigit() and len(right) == 3 and len(left.replace("-", "")) <= 3:

            try:

                return float(left.replace("-", "") + right) * (-1 if left.startswith("-") else 1)

            except ValueError:

                pass

    try:

        return float(comma_to_dot)

    except ValueError:

        return None





def _cell_float(val: Any) -> Optional[float]:

    if val is None:

        return None

    if isinstance(val, bool):

        return None

    if isinstance(val, (int, float)):

        return float(val)

    s = str(val).strip().replace(",", ".").replace(" ", "").replace("\u00a0", "")

    if not s:

        return None

    try:

        return float(s)

    except ValueError:

        return None





def merge_import_excel_overlay_into_product_data(

    product_data: Dict[str, Any],

    overlay: Optional[Dict[str, Any]],

) -> None:

    """Ghi đè shop + giá + SKU + tên TQ / shop TQ từ overlay Excel batch."""

    if not overlay or not isinstance(overlay, dict):

        _fill_shop_id_from_style_if_empty(product_data)

        return

    co = overlay.get("code")

    if co is not None and str(co).strip():

        product_data["code"] = str(co).strip().upper()

    if (sn := overlay.get("shop_name")) is not None and str(sn).strip():

        product_data["shop_name"] = str(sn).strip()

    pl = overlay.get("pro_lower_price")

    if pl is not None and str(pl).strip() != "":

        product_data["pro_lower_price"] = _cell_str(pl)

    ph = overlay.get("pro_high_price")

    if ph is not None and str(ph).strip() != "":

        product_data["pro_high_price"] = _cell_str(ph)

    num = _cell_float(overlay.get("price"))

    if num is not None:

        product_data["price"] = num

    if (sid := overlay.get("shop_id")) is not None and str(sid).strip():

        product_data["shop_id"] = str(sid).strip()

    if (sk := overlay.get("_sync_style_kieu_dang")) is not None and str(sk).strip():

        product_data["style"] = str(sk).strip()

    if (cn := overlay.get("chinese_name")) is not None and str(cn).strip():

        product_data["chinese_name"] = str(cn).strip()

    if (sc := overlay.get("shop_name_chinese")) is not None and str(sc).strip():

        product_data["shop_name_chinese"] = str(sc).strip()

    vnd_raw = overlay.get("price_display_vnd_g")

    if vnd_raw is not None and str(vnd_raw).strip() != "":

        _merge_excel_column_g_vnd_into_product_info(product_data, vnd_raw)

    _fill_shop_id_from_style_if_empty(product_data)





def _merge_excel_column_g_vnd_into_product_info(product_data: Dict[str, Any], raw: Any) -> None:

    pi_any = product_data.get("product_info")

    pi: Dict[str, Any]

    if isinstance(pi_any, dict):

        pi = pi_any

    else:

        pi = {}

        product_data["product_info"] = pi

    mk_any = pi.get("market_info")

    mk: Dict[str, Any] = mk_any if isinstance(mk_any, dict) else {}



    parsed = _parse_vnd_amount_from_excel_cell(raw)

    if parsed is not None:

        vn = round(parsed, 2)

        if abs(vn - round(vn)) < 1e-6:

            mk["price_vnd"] = int(round(vn))

        else:

            mk["price_vnd"] = vn

        mk.pop("price_vnd_display", None)

    else:

        s = _cell_str(raw)

        if not s:

            return

        mk["price_vnd_display"] = s

        mk.pop("price_vnd", None)



    mk.pop("note", None)

    mk["excel_price_vnd_source"] = "price_display_vnd_g (overlay legacy)"

    pi["market_info"] = mk





def _fill_shop_id_from_style_if_empty(product_data: Dict[str, Any]) -> None:

    cur = str(product_data.get("shop_id") or "").strip()

    if cur:

        return

    sty = product_data.get("style")

    if isinstance(sty, str) and sty.strip():

        product_data["shop_id"] = sty.strip()





def _labels_at(

    header_row1: Tuple[Any, ...], header_row2: Tuple[Any, ...], col_1based: int

) -> set[str]:

    out: set[str] = set()

    for row in (header_row1, header_row2):

        if 1 <= col_1based <= len(row):

            v = row[col_1based - 1]

            if v is not None and str(v).strip():

                out.add(_norm_header(v))

    return out





def _max_header_cols(header_row1: Tuple[Any, ...], header_row2: Tuple[Any, ...]) -> int:

    return max(len(header_row1), len(header_row2), 0)





def parse_link_import_excel(path: str | Path) -> Tuple[List[Dict[str, Any]], List[str]]:

    """

    Trả ([{excel_row, url, overlays}], skip_messages).



    Chỉ mẫu listing: **Link** + **China price / Giá Tệ** bắt buộc trong tiêu đề.

    Luôn đặt **`price`** (VNĐ) từ CN¥ × lưới × tỷ giá (không đọc giá VND từ Excel).

    """

    p = Path(path)

    skip: List[str] = []

    out: List[Dict[str, Any]] = []



    _header_tokens_only_link_like = frozenset(

        {

            _norm_header("link"),

            _norm_header("link sp"),

            _norm_header("url"),

            _norm_header("href"),

            _norm_header("id"),

        }

    )



    wb = load_workbook(p, read_only=True, data_only=True)

    try:

        ws = wb.active

        rows_hdr = ws.iter_rows(

            min_row=1,

            max_row=2,

            min_col=1,

            max_col=MAX_IMPORT_COL_1BASED,

            values_only=True,

        )

        rows_pair = list(rows_hdr)

        header = tuple(rows_pair[0]) if rows_pair else ()

        header2 = tuple(rows_pair[1] if len(rows_pair) > 1 else ())

        if not (

            any(x is not None and str(x).strip() for x in header)

            or any(x is not None and str(x).strip() for x in header2)

        ):

            return [], ["Dòng tiêu đề trống hoặc không đọc được."]



        mc = max(_max_header_cols(header, header2), MAX_IMPORT_COL_1BASED, len(header), len(header2))

        label_by_col = _merged_labels_for_cols(header, header2, mc)



        link_col = _resolve_link_column_required(label_by_col)

        china_price_col = _resolve_china_price_column_required(label_by_col)

        if link_col is None:

            return [], [

                "File không có cột Link trong tiêu đề (vd. «Link SP», item_url). "

                "Import batch chỉ nhận mẫu export listing tái nhập."

            ]

        if china_price_col is None:

            return [], [

                "File không có cột Giá Tệ / China price trong tiêu đề. "

                "Import batch chỉ nhận mẫu listing có giá nhân dân tệ."

            ]



        shop_cn_col = _resolve_shop_chinese_column(label_by_col)

        ch_name_col = _resolve_chinese_name_column(label_by_col)

        sku_hdr_col = _resolve_internal_sku_header_column(label_by_col)

        vnd_optional_col = _resolve_optional_vnd_per_cny_column(label_by_col)

        base_vnd_rate = _default_listing_import_vnd_per_cny()



        data_iter = ws.iter_rows(

            min_row=DATA_FIRST_ROW,

            max_row=DATA_FIRST_ROW + _MAX_ROWS,

            min_col=1,

            max_col=MAX_IMPORT_COL_1BASED,

            values_only=True,

        )

        for excel_row, row in enumerate(data_iter, start=DATA_FIRST_ROW):

            cells = tuple(row or ())

            if not any(v is not None and str(v).strip() != "" for v in cells):

                continue



            def coord(j: int) -> Optional[Any]:

                if j < 1 or j > len(cells):

                    return None

                return cells[j - 1]



            url = _cell_str(coord(link_col))



            if url and (_norm_header(url) in _header_tokens_only_link_like):

                continue



            if not _row_has_http_link(url):

                if url.strip() and (_norm_header(url) not in _header_tokens_only_link_like):

                    skip.append(f"Dòng {excel_row}: bỏ qua (thiếu link http hợp lệ).")

                continue



            overlays: Dict[str, Any] = {}

            if sku_hdr_col:

                sku_raw = _cell_str(coord(sku_hdr_col)).strip().upper()

                if sku_raw and INTERNAL_SKU_RE.fullmatch(sku_raw):

                    overlays["code"] = sku_raw



            if shop_cn_col:

                cn_shop_cell = coord(shop_cn_col)

                if cn_shop_cell is not None and str(cn_shop_cell).strip():

                    overlays["shop_name_chinese"] = _cell_str(cn_shop_cell)



            if ch_name_col:

                cn_name_cell = coord(ch_name_col)

                if cn_name_cell is not None and str(cn_name_cell).strip():

                    overlays["chinese_name"] = _cell_str(cn_name_cell)



            row_vnd_cell = coord(vnd_optional_col) if vnd_optional_col else None

            row_rate = _cell_float(row_vnd_cell)

            eff_vnd_per_cny = (

                row_rate if row_rate is not None and row_rate > 0 and math.isfinite(row_rate) else base_vnd_rate

            )



            cny_cell = coord(china_price_col)

            if cny_cell is None or str(cny_cell).strip() == "":

                skip.append(f"Dòng {excel_row}: thiếu Giá Tệ / China price.")

                continue



            overlays["pro_lower_price"] = _cell_str(cny_cell)

            parsed_cny = parse_approx_cny_amount_from_cell(cny_cell)

            if parsed_cny is None:

                skip.append(f"Dòng {excel_row}: không đọc được giá nhân dân tệ từ ô Giá Tệ.")

                continue



            coef = cny_exchange_multiplier_from_grid(parsed_cny)

            vnd_px = estimate_listing_vnd_rounded(parsed_cny, coef, eff_vnd_per_cny)

            if vnd_px is None:

                skip.append(f"Dòng {excel_row}: không quy đổi được Giá Tệ sang VNĐ.")

                continue

            overlays["price"] = float(vnd_px)



            out.append({"excel_row": excel_row, "url": url.strip(), "overlays": overlays})



    finally:

        wb.close()



    if not out and not skip:

        skip.append("Không có dòng dữ liệu có link sau dòng tiêu đề.")

    return out, skip


