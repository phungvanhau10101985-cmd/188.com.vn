"""Excel mẫu cho các import trên trang admin vận chuyển EMS."""

from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def _workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _bold_header_row(ws, row: int, values: list[str]) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True)


def build_ems_shipment_sample_xlsx() -> tuple[bytes, str]:
    """file gui ems.xlsx — cột A mã vận đơn, I đơn shop, G COD, D tên khách, H mã SP."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Gui EMS"
    headers = [
        "MA_VAN_DON",
        "TEN_SP",
        "TRONG_LUONG",
        "TEN_KH",
        "DIA_CHI_KH",
        "SDT_KH",
        "COD",
        "MA_SP",
        "DON_HANG",
    ]
    _bold_header_row(ws, 1, headers)
    ws.append(
        [
            "EE123456789VN",
            "Áo thun nam",
            0.5,
            "Nguyễn Văn A",
            "123 Đường ABC, Q.1, TP.HCM",
            "0901234567",
            150000,
            "H9441/1/xl",
            "DH131",
        ]
    )
    ws.append(
        [
            "EE987654321VN",
            "Quần jean",
            0.8,
            "Trần Thị B",
            "456 Đường XYZ, Q.3, TP.HCM",
            "0912345678",
            0,
            "H0723/40/3",
            "DC42",
        ]
    )
    note = (
        "Cột A: mã vận đơn EMS · I: mã đơn shop (DHxxx/DCxxx) · G: COD · D: tên khách · H: mã SP kho. "
        "Import lần 2: mã cột A đã có thì cập nhật, mã mới thì thêm dòng."
    )
    ws.cell(row=4, column=1, value=note)
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    return _workbook_to_bytes(wb), "file_gui_ems_mau.xlsx"


def build_cod_settlement_sample_xlsx() -> tuple[bytes, str]:
    """Doi soat cod — E1 ngày trả, từ hàng 3: B tham chiếu, C mã EMS, D số tiền."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Doi soat COD"
    ws.cell(row=1, column=5, value=date.today().strftime("%d/%m/%Y"))
    ws.cell(row=1, column=6, value="(E1: ngày EMS trả tiền cho shop)")
    _bold_header_row(ws, 2, ["", "Mã tham chiếu (B)", "Mã vận chuyển EMS (C)", "Số tiền đã trả (D)"])
    ws.append(["", "MA_THAM_CHIEU_01", "EE123456789VN", 5200000])
    ws.append(["", "", "EE987654321VN", 1500000])
    for col, width in [(2, 22), (3, 20), (4, 18), (5, 14)]:
        ws.column_dimensions[get_column_letter(col)].width = width
    return _workbook_to_bytes(wb), "doi_soat_cod_mau.xlsx"


def build_freight_settlement_sample_xlsx() -> tuple[bytes, str]:
    """Doi soat cuoc — A mã EMS, L cước phí."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Doi soat cuoc"
    _bold_header_row(ws, 1, ["MA_E1"])
    ws.cell(row=1, column=12, value="CUOC_PHI")
    ws.append(["EE123456789VN", None, None, None, None, None, None, None, None, None, None, 45000])
    ws.append(["EE987654321VN", None, None, None, None, None, None, None, None, None, None, 62000])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["L"].width = 14
    return _workbook_to_bytes(wb), "doi_soat_cuoc_mau.xlsx"


def build_shop_return_confirm_sample_xlsx() -> tuple[bytes, str]:
    """Xác nhận đơn hoàn — mỗi dòng một mã EMS / tham chiếu / DHxxx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Xac nhan hoan"
    _bold_header_row(ws, 1, ["Mã (EMS / tham chiếu / DHxxx)"])
    ws.append(["EE123456789VN"])
    ws.append(["MA_THAM_CHIEU_A"])
    ws.append(["DH131"])
    ws.column_dimensions["A"].width = 28
    return _workbook_to_bytes(wb), "xac_nhan_don_hoan_mau.xlsx"
